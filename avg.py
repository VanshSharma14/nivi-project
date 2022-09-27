from openpyxl import load_workbook
import utils

workbook = load_workbook(filename='data.xlsx')
ws = workbook.active

# print(ws['A12'].value)
index = 0
inArr = []
valArr = []
closest = 0
# it turns all those cells into an array
for row in ws.iter_rows(min_row=58, min_col=3, max_col=7, max_row=97, values_only=True):
        for cell in row:
                if cell != None:
                        valArr.append(cell)

print(valArr)
# this function gets the *closest to 300* values
# and adds them to a separate array so they can be used later
finalArr = utils.arr_index_finder(valArr)
